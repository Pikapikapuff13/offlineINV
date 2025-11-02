# -*- coding: utf-8 -*-
"""
Inventory Manager – Offline Windows Desktop App
Features:
- Supply goods to job workers
- Receive finished + damaged goods with unique attributes
- Dynamic remaining quantities
- Searchable tables, export PDF/Excel
- Autocomplete goods (add new on-the-fly)
- Dashboard with all views
"""

import sys
import os
import sqlite3
from datetime import datetime
from pathlib import Path

from PySide6.QtCore import (
    Qt, QDate, QSortFilterProxyModel, QAbstractTableModel,
    QModelIndex, QItemSelectionModel, Signal
)
from PySide6.QtGui import QIcon, QStandardItemModel, QStandardItem
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QPushButton, QLineEdit, QDateEdit, QSpinBox,
    QLabel, QTableView, QMessageBox, QComboBox, QFormLayout,
    QHeaderView, QToolBar, QStyleFactory, QFileDialog,
    QCompleter, QSpacerItem, QSizePolicy
)

# --------------------------------------------------------------
# Database Path
# --------------------------------------------------------------
DB_DIR = Path(__file__).with_name("data")
DB_DIR.mkdir(exist_ok=True)
DB_PATH = DB_DIR / "inventory.db"

# --------------------------------------------------------------
# Initialize SQLite Database
# --------------------------------------------------------------
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Goods master
    cur.execute("""CREATE TABLE IF NOT EXISTS goods (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL)""")

    # Supply
    cur.execute("""CREATE TABLE IF NOT EXISTS supply (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    invoice TEXT UNIQUE NOT NULL,
                    goods_id INTEGER NOT NULL,
                    qty INTEGER NOT NULL,
                    FOREIGN KEY(goods_id) REFERENCES goods(id))""")

    # Receipt
    cur.execute("""CREATE TABLE IF NOT EXISTS receipt (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    receipt_invoice TEXT UNIQUE NOT NULL,
                    supply_invoice TEXT NOT NULL,
                    goods_id INTEGER NOT NULL,
                    finished_qty INTEGER NOT NULL,
                    finished_attr TEXT,
                    damaged_qty INTEGER NOT NULL,
                    FOREIGN KEY(goods_id) REFERENCES goods(id),
                    FOREIGN KEY(supply_invoice) REFERENCES supply(invoice))""")

    # Auto-update remaining qty in supply
    cur.execute("""CREATE TRIGGER IF NOT EXISTS update_remaining
                   AFTER INSERT ON receipt
                   BEGIN
                       UPDATE supply
                       SET qty = qty - NEW.finished_qty - NEW.damaged_qty
                       WHERE invoice = NEW.supply_invoice;
                   END;""")

    conn.commit()
    conn.close()

# --------------------------------------------------------------
# Reusable SQL Table Model
# --------------------------------------------------------------
class SqlTableModel(QAbstractTableModel):
    def __init__(self, query, headers, parent=None):
        super().__init__(parent)
        self.headers = headers
        self.records = []
        self.refresh(query)

    def refresh(self, query):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute(query)
        self.records = cur.fetchall()
        conn.close()
        self.layoutAboutToBeChanged.emit()
        self.layoutChanged.emit()

    def rowCount(self, parent=QModelIndex()):
        return len(self.records)

    def columnCount(self, parent=QModelIndex()):
        return len(self.headers)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or role != Qt.DisplayRole:
            return None
        return str(self.records[index.row()][index.column()])

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.headers[section]
        return super().headerData(section, orientation, role)

# --------------------------------------------------------------
# Main Application Window
# --------------------------------------------------------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Inventory Manager")
        self.setWindowIcon(QIcon(resource_path("resources/icon.png")))
        self.resize(1150, 720)

        # Toolbar
        toolbar = QToolBar()
        self.addToolBar(toolbar)
        toolbar.addAction("Refresh All", self.refresh_all)

        # Tabs
        tabs = QTabWidget()
        self.setCentralWidget(tabs)

        tabs.addTab(self.build_dashboard(), "Dashboard")
        tabs.addTab(self.build_supply_tab(), "Supply")
        tabs.addTab(self.build_receipt_tab(), "Receipt")
        tabs.addTab(self.build_goods_tab(), "Goods")
        tabs.addTab(self.build_original_bill_tab(), "Original Supply Bill")
        tabs.addTab(self.build_dynamic_bill_tab(), "Dynamic Supply Bill")

        self.refresh_all()

    # ------------------------------------------------------------------
    # Dashboard
    # ------------------------------------------------------------------
    def build_dashboard(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.addWidget(QLabel("<h1>Inventory Dashboard</h1>"))

        grid = QHBoxLayout()
        for name in ["Supply", "Receipt", "Goods", "Original Bill", "Dynamic Bill"]:
            btn = QPushButton(name)
            btn.setFixedHeight(70)
            btn.clicked.connect(lambda _, n=name: self.parent().setCurrentWidget(
                self.parent().findChild(QWidget, objectName=n.replace(" ", ""))))
            grid.addWidget(btn)
        lay.addLayout(grid)
        lay.addStretch()
        return w

    # ------------------------------------------------------------------
    # Supply Tab
    # ------------------------------------------------------------------
    def build_supply_tab(self):
        w = QWidget()
        w.setObjectName("Supply")
        lay = QVBoxLayout(w)

        # Form
        form = QFormLayout()
        self.sup_date = QDateEdit(calendarPopup=True)
        self.sup_date.setDate(QDate.currentDate())
        self.sup_invoice = QLineEdit()
        self.sup_goods = QComboBox()
        self.sup_goods.setEditable(True)
        self.sup_goods.completer().setCompletionMode(QCompleter.PopupCompletion)
        self.sup_qty = QSpinBox()
        self.sup_qty.setRange(1, 999999)

        form.addRow("Date", self.sup_date)
        form.addRow("Invoice #", self.sup_invoice)
        form.addRow("Goods", self.sup_goods)
        form.addRow("Quantity", self.sup_qty)

        btn_add = QPushButton("Add Supply")
        btn_add.clicked.connect(self.add_supply)
        form.addRow(btn_add)
        lay.addLayout(form)

        # Table
        self.sup_model = SqlTableModel(
            """SELECT s.date, s.invoice, g.name, s.qty
               FROM supply s JOIN goods g ON s.goods_id=g.id
               ORDER BY s.date DESC""",
            ["Date", "Invoice", "Goods", "Qty"]
        )
        proxy = QSortFilterProxyModel()
        proxy.setSourceModel(self.sup_model)
        self.sup_view = QTableView()
        self.sup_view.setModel(proxy)
        self.sup_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.sup_view.setSortingEnabled(True)

        search = QLineEdit()
        search.setPlaceholderText("Search…")
        search.textChanged.connect(proxy.setFilterFixedString)
        lay.addWidget(search)
        lay.addWidget(self.sup_view)

        # Export
        exp = QHBoxLayout()
        exp.addWidget(QPushButton("Export PDF", clicked=lambda: self.export_pdf(self.sup_view, "Supply")))
        exp.addWidget(QPushButton("Export Excel", clicked=lambda: self.export_excel(self.sup_view, "Supply")))
        lay.addLayout(exp)

        self.load_goods_combobox()
        return w

    # ------------------------------------------------------------------
    # Receipt Tab
    # ------------------------------------------------------------------
    def build_receipt_tab(self):
        w = QWidget()
        w.setObjectName("Receipt")
        lay = QVBoxLayout(w)

        form = QFormLayout()
        self.rec_date = QDateEdit(calendarPopup=True)
        self.rec_date.setDate(QDate.currentDate())
        self.rec_invoice = QLineEdit()
        self.rec_supply_invoice = QComboBox()
        self.rec_goods = QLineEdit()
        self.rec_goods.setReadOnly(True)
        self.rec_finished_qty = QSpinBox()
        self.rec_finished_attr = QLineEdit()
        self.rec_damaged_qty = QSpinBox()

        form.addRow("Date", self.rec_date)
        form.addRow("Receipt #", self.rec_invoice)
        form.addRow("Supply Invoice", self.rec_supply_invoice)
        form.addRow("Goods", self.rec_goods)
        form.addRow("Finished Qty", self.rec_finished_qty)
        form.addRow("Unique Attribute", self.rec_finished_attr)
        form.addRow("Damaged Qty", self.rec_damaged_qty)

        btn_add = QPushButton("Add Receipt")
        btn_add.clicked.connect(self.add_receipt)
        form.addRow(btn_add)
        lay.addLayout(form)

        # Table
        self.rec_model = SqlTableModel(
            """SELECT r.date, r.receipt_invoice, r.supply_invoice,
                      g.name, r.finished_qty, r.finished_attr, r.damaged_qty
               FROM receipt r JOIN goods g ON r.goods_id=g.id
               ORDER BY r.date DESC""",
            ["Date", "Receipt#", "Supply#", "Goods", "Finished", "Attribute", "Damaged"]
        )
        proxy = QSortFilterProxyModel()
        proxy.setSourceModel(self.rec_model)
        self.rec_view = QTableView()
        self.rec_view.setModel(proxy)
        self.rec_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        search = QLineEdit()
        search.setPlaceholderText("Search…")
        search.textChanged.connect(proxy.setFilterFixedString)
        lay.addWidget(search)
        lay.addWidget(self.rec_view)

        # Export
        exp = QHBoxLayout()
        exp.addWidget(QPushButton("Export PDF", clicked=lambda: self.export_pdf(self.rec_view, "Receipt")))
        exp.addWidget(QPushButton("Export Excel", clicked=lambda: self.export_excel(self.rec_view, "Receipt")))
        lay.addLayout(exp)

        self.rec_supply_invoice.currentTextChanged.connect(self.load_supply_goods)
        return w

    # ------------------------------------------------------------------
    # Goods Master Tab
    # ------------------------------------------------------------------
    def build_goods_tab(self):
        w = QWidget()
        w.setObjectName("Goods")
        lay = QVBoxLayout(w)

        add_lay = QHBoxLayout()
        self.new_goods = QLineEdit()
        btn = QPushButton("Add Goods")
        btn.clicked.connect(self.add_goods)
        add_lay.addWidget(QLabel("New Goods:"))
        add_lay.addWidget(self.new_goods)
        add_lay.addWidget(btn)
        lay.addLayout(add_lay)

        self.goods_model = SqlTableModel("SELECT name FROM goods ORDER BY name", ["Goods Name"])
        proxy = QSortFilterProxyModel()
        proxy.setSourceModel(self.goods_model)
        view = QTableView()
        view.setModel(proxy)
        view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        search = QLineEdit()
        search.setPlaceholderText("Search…")
        search.textChanged.connect(proxy.setFilterFixedString)
        lay.addWidget(search)
        lay.addWidget(view)
        return w

    # ------------------------------------------------------------------
    # Original Bill Tab
    # ------------------------------------------------------------------
    def build_original_bill_tab(self):
        w = QWidget()
        w.setObjectName("OriginalSupplyBill")
        lay = QVBoxLayout(w)

        top = QHBoxLayout()
        self.orig_invoice = QComboBox()
        btn = QPushButton("Show")
        btn.clicked.connect(self.show_original_bill)
        top.addWidget(QLabel("Invoice:"))
        top.addWidget(self.orig_invoice)
        top.addWidget(btn)
        lay.addLayout(top)

        self.orig_view = QTableView()
        lay.addWidget(self.orig_view)

        exp = QHBoxLayout()
        exp.addWidget(QPushButton("Export PDF", clicked=lambda: self.export_pdf(self.orig_view, "Original")))
        exp.addWidget(QPushButton("Export Excel", clicked=lambda: self.export_excel(self.orig_view, "Original")))
        lay.addLayout(exp)
        return w

    # ------------------------------------------------------------------
    # Dynamic Bill Tab
    # ------------------------------------------------------------------
    def build_dynamic_bill_tab(self):
        w = QWidget()
        w.setObjectName("DynamicSupplyBill")
        lay = QVBoxLayout(w)

        top = QHBoxLayout()
        self.dyn_invoice = QComboBox()
        btn = QPushButton("Show")
        btn.clicked.connect(self.show_dynamic_bill)
        top.addWidget(QLabel("Invoice:"))
        top.addWidget(self.dyn_invoice)
        top.addWidget(btn)
        lay.addLayout(top)

        self.dyn_view = QTableView()
        lay.addWidget(self.dyn_view)

        exp = QHBoxLayout()
        exp.addWidget(QPushButton("Export PDF", clicked=lambda: self.export_pdf(self.dyn_view, "Dynamic")))
        exp.addWidget(QPushButton("Export Excel", clicked=lambda: self.export_excel(self.dyn_view, "Dynamic")))
        lay.addLayout(exp)
        return w

    # ------------------------------------------------------------------
    # Helper Methods
    # ------------------------------------------------------------------
    def refresh_all(self):
        self.load_goods_combobox()
        self.load_supply_invoices()
        self.load_open_supply_invoices()
        self.sup_model.refresh(self.sup_model.query().lastQuery())
        self.rec_model.refresh(self.rec_model.query().lastQuery())
        self.goods_model.refresh(self.goods_model.query().lastQuery())

    def load_goods_combobox(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT name FROM goods ORDER BY name")
        names = [r[0] for r in cur.fetchall()]
        conn.close()

        self.sup_goods.clear()
        self.sup_goods.addItems(names)
        completer = QCompleter(names)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.sup_goods.setCompleter(completer)

    def add_goods(self):
        name = self.new_goods.text().strip()
        if not name:
            QMessageBox.warning(self, "Error", "Enter goods name")
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            cur.execute("INSERT INTO goods (name) VALUES (?)", (name,))
            conn.commit()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Error", "Goods already exists")
        finally:
            conn.close()
        self.new_goods.clear()
        self.refresh_all()

    def add_supply(self):
        date = self.sup_date.date().toString("yyyy-MM-dd")
        invoice = self.sup_invoice.text().strip()
        goods_name = self.sup_goods.currentText().strip()
        qty = self.sup_qty.value()

        if not (invoice and goods_name and qty):
            QMessageBox.warning(self, "Error", "Fill all fields")
            return

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT id FROM goods WHERE name=?", (goods_name,))
        row = cur.fetchone()
        if not row:
            cur.execute("INSERT INTO goods (name) VALUES (?)", (goods_name,))
            goods_id = cur.lastrowid
        else:
            goods_id = row[0]

        try:
            cur.execute("INSERT INTO supply (date, invoice, goods_id, qty) VALUES (?,?,?,?)",
                        (date, invoice, goods_id, qty))
            conn.commit()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Error", "Invoice already used")
            conn.close()
            return
        conn.close()

        self.sup_invoice.clear()
        self.sup_qty.setValue(1)
        self.refresh_all()

    def load_open_supply_invoices(self):
        self.rec_supply_invoice.clear()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("""SELECT DISTINCT s.invoice
                       FROM supply s
                       WHERE NOT EXISTS (
                           SELECT 1 FROM receipt r
                           WHERE r.supply_invoice = s.invoice
                           AND r.finished_qty + r.damaged_qty >= s.qty
                       )
                       ORDER BY s.invoice""")
        for (inv,) in cur.fetchall():
            self.rec_supply_invoice.addItem(inv)
        conn.close()

    def load_supply_goods(self, invoice):
        if not invoice:
            self.rec_goods.clear()
            return
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("""SELECT g.name, s.qty,
                              COALESCE(SUM(r.finished_qty + r.damaged_qty), 0)
                       FROM supply s
                       JOIN goods g ON s.goods_id = g.id
                       LEFT JOIN receipt r ON s.invoice = r.supply_invoice
                       WHERE s.invoice = ?
                       GROUP BY s.id""", (invoice,))
        row = cur.fetchone()
        conn.close()
        if row:
            name, total, used = row
            self.rec_goods.setText(name)
            remain = total - used
            self.rec_finished_qty.setMaximum(remain)
            self.rec_damaged_qty.setMaximum(remain)

    def add_receipt(self):
        date = self.rec_date.date().toString("yyyy-MM-dd")
        rec_inv = self.rec_invoice.text().strip()
        sup_inv = self.rec_supply_invoice.currentText()
        fin_qty = self.rec_finished_qty.value()
        attr = self.rec_finished_attr.text().strip()
        dam_qty = self.rec_damaged_qty.value()

        if not (rec_inv and sup_inv and (fin_qty + dam_qty > 0)):
            QMessageBox.warning(self, "Error", "Fill required fields")
            return

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT goods_id FROM supply WHERE invoice=?", (sup_inv,))
        goods_id = cur.fetchone()[0]

        try:
            cur.execute("""INSERT INTO receipt
                           (date, receipt_invoice, supply_invoice, goods_id,
                            finished_qty, finished_attr, damaged_qty)
                           VALUES (?,?,?,?,?,?,?)""",
                        (date, rec_inv, sup_inv, goods_id, fin_qty, attr, dam_qty))
            conn.commit()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Error", "Receipt invoice exists")
            conn.close()
            return
        conn.close()

        self.rec_invoice.clear()
        self.rec_finished_qty.setValue(0)
        self.rec_damaged_qty.setValue(0)
        self.rec_finished_attr.clear()
        self.refresh_all()

    def load_supply_invoices(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT invoice FROM supply ORDER BY date DESC")
        invs = [r[0] for r in cur.fetchall()]
        conn.close()
        self.orig_invoice.clear()
        self.orig_invoice.addItems(invs)
        self.dyn_invoice.clear()
        self.dyn_invoice.addItems(invs)

    def show_original_bill(self):
        inv = self.orig_invoice.currentText()
        if not inv:
            return
        query = f"""SELECT s.date, s.invoice, g.name, s.qty
                    FROM supply s JOIN goods g ON s.goods_id=g.id
                    WHERE s.invoice='{inv}'"""
        model = SqlTableModel(query, ["Date", "Invoice", "Goods", "Qty"])
        self.orig_view.setModel(model)
        self.orig_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def show_dynamic_bill(self):
        inv = self.dyn_invoice.currentText()
        if not inv:
            return
        query = f"""SELECT s.date, s.invoice, g.name, s.qty,
                           COALESCE(SUM(r.finished_qty),0) AS fin,
                           COALESCE(SUM(r.damaged_qty),0) AS dam,
                           s.qty - COALESCE(SUM(r.finished_qty),0) - COALESCE(SUM(r.damaged_qty),0) AS remain
                    FROM supply s
                    JOIN goods g ON s.goods_id=g.id
                    LEFT JOIN receipt r ON s.invoice=r.supply_invoice
                    WHERE s.invoice='{inv}'
                    GROUP BY s.id"""
        model = SqlTableModel(query,
            ["Date", "Invoice", "Goods", "Supplied", "Finished", "Damaged", "Remaining"])
        self.dyn_view.setModel(model)
        self.dyn_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def export_pdf(self, view, name):
        path, _ = QFileDialog.getSaveFileName(self, "Save PDF", f"{name}.pdf", "PDF (*.pdf)")
        if not path:
            return
        from PySide6.QtPrintSupport import QPrinter
        from PySide6.QtGui import QPainter
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPrinter.A4)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(path)
        painter = QPainter(printer)
        view.render(painter)
        painter.end()
        QMessageBox.information(self, "Done", f"Saved: {path}")

    def export_excel(self, view, name):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", f"{name}.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = name
            model = view.model()
            for c in range(model.columnCount()):
                ws.cell(1, c+1, model.headerData(c, Qt.Horizontal))
            for r in range(model.rowCount()):
                for c in range(model.columnCount()):
                    idx = model.index(r, c)
                    ws.cell(r+2, c+1, model.data(idx))
            wb.save(path)
            QMessageBox.information(self, "Done", f"Saved: {path}")
        except ImportError:
            QMessageBox.critical(self, "Error", "Install openpyxl: pip install openpyxl")

# --------------------------------------------------------------
# Resource Path (for .exe)
# --------------------------------------------------------------
def resource_path(relative_path):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# --------------------------------------------------------------
# Run App
# --------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    win = MainWindow()
    win.show()
    sys.exit(app.exec())